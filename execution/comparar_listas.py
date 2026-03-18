"""
Compara duas listas de preços em Excel.

Suporta dois formatos:
  – "Custom Tabela de Preços": header dinâmico (detectado por conteúdo),
    colunas 'Código', 'Descrição', 'Preço REVENDA (1 a 5 unid.)', etc.
  – "Lista Mestra / padrão": colunas 'PN', 'DESCRIÇÃO', '1 - 5', '6 - 10', '10+'

Retorna dict com:
    identical       : bool  — True quando não há diferença alguma
    new_count       : int   — Produtos presentes apenas na lista nova
    removed_count   : int   — Produtos presentes apenas na lista antiga (removidos)
    modified_count  : int   — Produtos com alteração de preço
    unchanged_count : int   — Produtos idênticos
    total_new_file  : int   — Total de linhas de produto na lista nova
    total_old_file  : int   — Total de linhas de produto na lista antiga
    col_info        : dict  — Colunas detectadas (para debug)
    excel_bytes     : bytes — Apenas quando identical=False
"""
import io
import logging


import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz

log = logging.getLogger("comparador.engine")

# ─────────────────────── Estilos ────────────────────────────────
_FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
_FILL_NEW     = PatternFill("solid", fgColor="C6EFCE")
_FILL_INC     = PatternFill("solid", fgColor="FFC7CE")
_FILL_DEC     = PatternFill("solid", fgColor="FFEB9C")
_FILL_REMOVED = PatternFill("solid", fgColor="D9D9D9")  # Cinza — produto removido

_FONT_HEADER  = Font(bold=True, color="FFFFFF", size=11)
_FONT_NEW     = Font(bold=True, color="375623")
_FONT_INC     = Font(bold=True, color="9C0006")
_FONT_DEC     = Font(bold=True, color="7D6608")
_FONT_REMOVED = Font(italic=True, color="595959")
_FONT_STATUS  = Font(bold=True)

# Faixas de preço: label canônico → candidatos de substring
_PRICE_LABELS = ["1-5", "5-10", "10+"]
_PRICE_CANDIDATES = {
    "1-5":  ["1 A 5", "1-5", "1 - 5", "1A5", "(1 A 5", "(1-5"],
    "5-10": ["6 A 10", "5-10", "6 - 10", "5 - 10", "6A10", "(6 A 10", "(5-10", "(6-10"],
    "10+":  [">10", "10+", "10 +", "> 10", "MAIS DE 10", "(>10"],
}

# Divisor para calcular DISTRI a partir do preço 1-5
_DISTRI_DIVISOR = 1.23

# Limites do double-check (ratio = preço_nova / DISTRI_NOVA)
# Vermelho se ratio > max OU ratio < min
_DC_LIMITS = {
    "DC 1-5":  {"max": 1.24, "min": 1.22},
    "DC 6-10": {"max": 1.21, "min": 1.19},
    "DC 10+":  {"max": 1.18, "min": 1.16},
}
_DC_LABELS = ["DC 1-5", "DC 6-10", "DC 10+"]

# Estilos extra para alerta do double-check
_FILL_DC_ALERT = PatternFill("solid", fgColor="FF0000")   # vermelho: fora dos limites
_FONT_DC_ALERT = Font(bold=True, color="FFFFFF")


# ═══════════════════════════════════════════════════════════════
# Leitura inteligente do Excel
# ═══════════════════════════════════════════════════════════════
def _read_excel_smart(file_bytes: bytes, sheet_name: str | int = 0) -> pd.DataFrame:
    """
    Lê o Excel detectando automaticamente onde está o header real.
    Suporta:
      1. Header na primeira linha (padrão)
      2. Header embutido algumas linhas abaixo (como no Custom Tabela)
    Remove linhas de 'seção/categoria' que não têm PN/Código preenchido.
    """
    log.debug("Lendo Excel — sheet=%r", sheet_name)
    raw = pd.read_excel(io.BytesIO(file_bytes), header=None, dtype=str,
                        sheet_name=sheet_name)
    log.debug("Arquivo carregado — %d linhas brutas, %d colunas", len(raw), len(raw.columns))

    # Procura a linha que contém os rótulos reais das colunas
    header_row_idx = _find_header_row(raw)
    log.debug("Header detectado na linha %d", header_row_idx)

    # Recarrega usando a linha correta como header
    df = pd.read_excel(
        io.BytesIO(file_bytes),
        header=header_row_idx,
        dtype=str,
        sheet_name=sheet_name,
    )
    df.columns = [str(c).strip() for c in df.columns]
    log.debug("Colunas após recarregamento: %s", list(df.columns))

    # Remove linhas completamente vazias ou que são "repetição de header"
    # (acontece no Custom Tabela onde cada seção repete o cabeçalho)
    header_signature = _make_header_signature(df.columns)
    mask_repeated = df.apply(
        lambda row: _make_header_signature(row.astype(str).str.strip().tolist()) == header_signature,
        axis=1,
    )
    n_repeated = mask_repeated.sum()
    if n_repeated:
        log.debug("Removidas %d linhas de header duplicado", n_repeated)
    df = df[~mask_repeated].copy()
    df = df.dropna(how="all").reset_index(drop=True)
    log.debug("%d linhas após limpeza", len(df))

    return df


def _find_header_row(raw: pd.DataFrame) -> int:
    """
    Percorre as primeiras linhas do DataFrame bruto procurando aquela que
    contém palavras-chave identificadoras do header real.
    """
    HEADER_KEYWORDS = {
        "PN", "CÓDIGO", "CODIGO", "CODE", "COD",
        "DESCRIÇÃO", "DESCRICAO", "DESCRIPTION",
        "1 A 5", "1-5", "1 - 5",
    }
    for idx, row in raw.iterrows():
        vals = {str(v).strip().upper() for v in row.values if pd.notna(v) and str(v).strip()}
        if vals & HEADER_KEYWORDS:
            return idx
        if idx > 20:  # Não procura além das primeiras 20 linhas
            break
    return 0  # Fallback: primeira linha


def _make_header_signature(values: list) -> frozenset:
    return frozenset(str(v).strip().upper()[:20] for v in values if str(v).strip())


# ═══════════════════════════════════════════════════════════════
# Detecção de colunas
# ═══════════════════════════════════════════════════════════════
def _find_col(df: pd.DataFrame, candidates: list[str]) -> str | None:
    """Encontra coluna cujo nome (upper) contém qualquer candidato."""
    upper_map = {str(c).upper(): c for c in df.columns}
    for cand in candidates:
        cand_up = cand.upper()
        for key, actual in upper_map.items():
            if cand_up in key:
                return actual
    return None


def _find_price_col(df: pd.DataFrame, label: str) -> str | None:
    """Detecta coluna de preço usando os candidatos mapeados para cada faixa."""
    return _find_col(df, _PRICE_CANDIDATES.get(label, [label]))


def _detect_columns(df: pd.DataFrame) -> dict:
    """Retorna dict com as colunas detectadas para PN, descrição e preços."""
    pn_candidates   = ["PN", "CÓDIGO", "CODIGO", "CODE", "COD", "PART NUMBER", "PART_NUMBER"]
    desc_candidates = ["DESCRIÇÃO", "DESCRICAO", "DESCRIPTION", "NOME", "PRODUTO", "DESC"]

    result = {
        "pn":   _find_col(df, pn_candidates),
        "desc": _find_col(df, desc_candidates),
    }
    for lbl in _PRICE_LABELS:
        result[lbl] = _find_price_col(df, lbl)
    log.debug("Colunas detectadas: %s", result)
    return result


# ═══════════════════════════════════════════════════════════════
# Utilitários numéricos
# ═══════════════════════════════════════════════════════════════
def _to_float(val) -> float | None:
    """Converte para float; lida com R$, vírgula decimal, etc."""
    try:
        v = float(
            str(val)
            .replace(",", ".")
            .replace("R$", "")
            .replace("\xa0", "")
            .replace(" ", "")
        )
        return None if np.isnan(v) else v
    except Exception:
        return None


def _pct_diff(old_val, new_val) -> float | None:
    """
    Diferença percentual como decimal: 0.0523 = +5,23%.
    Retorna None se não calculável.
    """
    old = _to_float(old_val)
    new = _to_float(new_val)
    if old is None or new is None or old == 0:
        return None
    return round((new - old) / old, 6)


def _norm(s) -> str:
    return str(s).strip().upper()


def _is_product_row(row: pd.Series, pn_col: str) -> bool:
    """Filtra linhas de seção/categoria: exige que o PN seja não-vazio."""
    val = str(row.get(pn_col, "")).strip()
    return bool(val) and val.upper() not in {"NAN", "NONE", "PN", "CÓDIGO", "CODIGO", "COD"}


# ═══════════════════════════════════════════════════════════════
# Comparação principal
# ═══════════════════════════════════════════════════════════════
def compare_excels(
    old_bytes: bytes,
    new_bytes: bytes,
    old_sheet: str | int = 0,
    new_sheet: str | int = 0,
) -> dict:
    log.info("compare_excels iniciado — old_sheet=%r  new_sheet=%r", old_sheet, new_sheet)

    old_df = _read_excel_smart(old_bytes, sheet_name=old_sheet)
    new_df = _read_excel_smart(new_bytes, sheet_name=new_sheet)

    old_cols = _detect_columns(old_df)
    new_cols = _detect_columns(new_df)
    log.info("Colunas OLD=%s", old_cols)
    log.info("Colunas NEW=%s", new_cols)

    if old_cols["pn"] is None:
        raise ValueError(
            "Coluna de PN/Código não encontrada na lista ANTIGA.\n"
            f"Colunas disponíveis: {list(old_df.columns)}"
        )
    if new_cols["pn"] is None:
        raise ValueError(
            "Coluna de PN/Código não encontrada na lista NOVA.\n"
            f"Colunas disponíveis: {list(new_df.columns)}"
        )

    # Filtrar apenas linhas de produto real
    old_df = old_df[old_df.apply(lambda r: _is_product_row(r, old_cols["pn"]), axis=1)].copy()
    new_df = new_df[new_df.apply(lambda r: _is_product_row(r, new_cols["pn"]), axis=1)].copy()
    log.info("Produtos filtrados — OLD=%d  NEW=%d", len(old_df), len(new_df))
    old_df = old_df.reset_index(drop=True)
    new_df = new_df.reset_index(drop=True)

    # Lookup da lista antiga: PN normalizado → Series
    old_lookup: dict[str, pd.Series] = {
        _norm(str(row[old_cols["pn"]])): row
        for _, row in old_df.iterrows()
    }

    # Rastreia quais PNs da lista antiga foram encontrados na nova
    matched_old_pns: set[str] = set()

    row_info: list[dict] = []

    for df_idx, new_row in new_df.iterrows():
        new_pn   = _norm(str(new_row[new_cols["pn"]]))
        new_desc = _norm(str(new_row.get(new_cols["desc"], ""))) if new_cols["desc"] else ""

        # 1. Match exato por PN
        old_row    = old_lookup.get(new_pn)
        match_type = "PN exato"
        if old_row is not None:
            matched_old_pns.add(new_pn)

        # 2. Match fuzzy por PN (≥ 85 %)
        if old_row is None:
            best_score, best_key = 0, None
            for key in old_lookup:
                s = fuzz.ratio(new_pn, key)
                if s > best_score:
                    best_score, best_key = s, key
            if best_score >= 85:
                old_row    = old_lookup[best_key]
                match_type = f"PN similar ({best_score}%)"
                matched_old_pns.add(best_key)
                log.debug("Fuzzy PN: %r → %r  score=%d", new_pn, best_key, best_score)

        # 3. Match fuzzy por descrição (≥ 80 %)
        if old_row is None and old_cols["desc"] and new_cols["desc"]:
            best_score, best_row = 0, None
            for _, orow in old_df.iterrows():
                old_desc = _norm(str(orow.get(old_cols["desc"], "")))
                s = fuzz.partial_ratio(new_desc, old_desc)
                if s > best_score:
                    best_score, best_row = s, orow
            if best_score >= 80:
                old_row    = best_row
                match_type = f"Descrição similar ({best_score}%)"
                matched_old_pns.add(_norm(str(best_row[old_cols["pn"]])))

        # 4. Produto novo
        if old_row is None:
            log.debug("NOVO produto: PN=%r  DESC=%r", new_pn, new_desc[:60])
            row_info.append({
                "idx": df_idx, "status": "NOVO",
                "pct": {lbl: None for lbl in _PRICE_LABELS},
                "old_prices": {lbl: None for lbl in _PRICE_LABELS},
                "match_type": "—",
            })
            continue

        # Comparar preços e guardar valores absolutos da lista antiga
        pct        = {}
        old_prices = {}
        diff       = False
        for lbl in _PRICE_LABELS:
            nc = new_cols.get(lbl)
            oc = old_cols.get(lbl)
            if nc and oc:
                v = _pct_diff(old_row.get(oc), new_row.get(nc))
                pct[lbl]        = v
                old_prices[lbl] = _to_float(old_row.get(oc))
                if v is not None and abs(v) > 1e-6:
                    diff = True
            else:
                pct[lbl]        = None
                old_prices[lbl] = None

        row_info.append({
            "idx": df_idx,
            "status": "ALTERADO" if diff else "IGUAL",
            "pct": pct,
            "old_prices": old_prices,
            "match_type": match_type,
        })

    new_count       = sum(1 for r in row_info if r["status"] == "NOVO")
    modified_count  = sum(1 for r in row_info if r["status"] == "ALTERADO")
    unchanged_count = sum(1 for r in row_info if r["status"] == "IGUAL")

    # Produtos removidos: existiam na lista antiga mas não foram matcheados na nova
    removed_df = old_df[
        old_df.apply(
            lambda r: _norm(str(r[old_cols["pn"]])) not in matched_old_pns,
            axis=1,
        )
    ].copy().reset_index(drop=True)
    removed_count = len(removed_df)

    # Renomear colunas do removed_df para os nomes equivalentes na lista nova,
    # usando old_cols/new_cols como dicionário de tradução (chave semântica = pivot).
    # Ex: old_cols["pn"] = "Código"  → new_cols["pn"] = "Código Produto"
    # Assim as colunas do removed_df ficam alinhadas com final_cols do Excel.
    if not removed_df.empty:
        col_rename = {}
        for key in list(old_cols.keys()) + list(_PRICE_LABELS):
            # chave semântica pode estar em old_cols ou direto nos labels de preço
            old_name = old_cols.get(key)
            new_name = new_cols.get(key)
            if old_name and new_name and old_name != new_name and old_name in removed_df.columns:
                col_rename[old_name] = new_name
        if col_rename:
            log.debug("Renomeando colunas do removed_df: %s", col_rename)
            removed_df = removed_df.rename(columns=col_rename)

    log.info(
        "Resumo — novos=%d  removidos=%d  alterados=%d  iguais=%d  total_new=%d  total_old=%d",
        new_count, removed_count, modified_count, unchanged_count, len(new_df), len(old_df),
    )
    if removed_count:
        for _, rrow in removed_df.iterrows():
            pn   = rrow.get(old_cols["pn"], "")
            desc = rrow.get(old_cols["desc"], "") if old_cols["desc"] else ""
            log.debug("REMOVIDO: PN=%r  DESC=%r", pn, str(desc)[:60])

    has_diff = new_count > 0 or removed_count > 0 or modified_count > 0

    result = {
        "identical":       not has_diff,
        "new_count":       new_count,
        "removed_count":   removed_count,
        "modified_count":  modified_count,
        "unchanged_count": unchanged_count,
        "total_new_file":  len(new_df),
        "total_old_file":  len(old_df),
        "col_info": {
            "old": old_cols,
            "new": new_cols,
        },
    }

    if not result["identical"]:
        result["excel_bytes"] = _build_excel(
            new_df, new_cols, row_info,
            removed_df=removed_df, old_cols=old_cols,
        )

    return result


# ═══════════════════════════════════════════════════════════════
# Geração do Excel de saída
# ═══════════════════════════════════════════════════════════════
def _build_excel(
    new_df: pd.DataFrame,
    new_cols: dict,
    row_info: list,
    removed_df: pd.DataFrame | None = None,
    old_cols: dict | None = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparação de Preços"

    status_map = {r["idx"]: r for r in row_info}
    col_names  = list(new_df.columns)

    # Coluna de preço da lista nova (1-5) — base para DISTRI
    price_15_col_new = new_cols.get("1-5")

    # Inserir colunas extras após a última coluna de preço (10+)
    ten_plus_col = new_cols.get("10+")
    insert_after = (
        col_names.index(ten_plus_col)
        if ten_plus_col and ten_plus_col in col_names
        else len(col_names) - 1
    )

    # Prefixos para deixar claro quais preços são da lista NOVA e quais da ANTIGA
    # Renomear colunas de preço da lista nova para prefixar "NOVA"
    # Mapeamento label canônico → nome de coluna curto nas 3 faixas
    # Ex: "1-5" → "NOVA 1-5" / "6-10" → "NOVA 6-10" / "10+" → "NOVA 10+"
    # (os dados reais continuam lidos pela coluna orig detectada em new_cols)
    _price_short = {"1-5": "NOVA 1-5", "5-10": "NOVA 6-10", "10+": "NOVA 10+"}
    _antiga_short = {"1-5": "ANTIGA 1-5", "5-10": "ANTIGA 6-10", "10+": "ANTIGA 10+"}

    # Colunas de preço da lista nova que serão substituídas pelos nomes curtos no cabeçalho
    price_orig_to_short: dict[str, str] = {}   # orig_col_name → "NOVA X"
    for lbl in _PRICE_LABELS:
        orig = new_cols.get(lbl)
        if orig and orig in col_names:
            price_orig_to_short[orig] = _price_short[lbl]

    # Nomes a exibir no cabeçalho (colunas originais trocadas pelos short labels)
    display_col_names = [price_orig_to_short.get(c, c) for c in col_names]

    # Cabeçalhos para as colunas extras (ANTIGA, %, DISTRI, DC)
    antiga_col_headers = [
        _antiga_short[lbl] for lbl in _PRICE_LABELS if new_cols.get(lbl)
    ]

    extra_after = (
        antiga_col_headers
        + [f"% {lbl}" for lbl in _PRICE_LABELS]
        + ["DISTRI NOVA"]
        + _DC_LABELS
        + ["DISTRI ANTIGA", "DISTRI %"]
    )

    # Nomes reservados pelas colunas extras calculadas + STATUS.
    # Se o arquivo de entrada for uma saída anterior do sistema, ele já tem essas
    # colunas; excluímos da parte "tail" de display_col_names para evitar
    # duplicatas em final_cols (que causariam col_idx a apontar para o lugar errado).
    _reserved = set(extra_after) | {"STATUS"} | set(_price_short.values()) | set(_antiga_short.values())

    def _is_reserved_col(disp_name: str) -> bool:
        """Retorna True se a coluna de display deve ser omitida do tail (é calculada/reservada)."""
        if disp_name in _reserved:
            return True
        # Colunas como "ANTIGA Preço REVENDA (1 a 5 unid.)" que vêm de saídas anteriores
        if disp_name.startswith("ANTIGA "):
            return True
        # Rótulos DC soltos ("1-5", "6-10", "10+") que surgem em saídas anteriores como colunas extras
        if disp_name in {"1-5", "6-10", "10+", "DC 1-5", "DC 6-10", "DC 10+"}:
            return True
        return False


    # Também precisamos filtrar os pares (orig_name, display_name) para não
    # re-escrever pelo loop de colunas não-preço nas células das extras calculadas.
    tail_pairs = [
        (orig, disp)
        for orig, disp in zip(col_names[insert_after + 1:], display_col_names[insert_after + 1:])
        if not _is_reserved_col(disp)
    ]

    tail_col_names     = [p[0] for p in tail_pairs]
    tail_display_names = [p[1] for p in tail_pairs]

    final_cols = (
        display_col_names[: insert_after + 1]
        + extra_after
        + tail_display_names
        + ["STATUS"]
    )

    # Atualizar referências que usam os short labels
    nova_price_display = set(_price_short.values())


    # Mapa coluna → índice (1-based)
    col_idx = {name: i + 1 for i, name in enumerate(final_cols)}

    # ── Cabeçalho ──────────────────────────────────────────────────
    # Cores de cabeçalho diferenciadas por grupo
    _FILL_HDR_NOVA   = PatternFill("solid", fgColor="1F4E79")  # azul escuro — preço nova
    _FILL_HDR_ANTIGA = PatternFill("solid", fgColor="375623")  # verde escuro — preço antiga
    _FILL_HDR_PCT    = PatternFill("solid", fgColor="7B2C2C")  # vinho — variação %
    _FILL_HDR_DISTRI = PatternFill("solid", fgColor="5B4C8A")  # roxo — DISTRI / double-check
    _FILL_HDR_STATUS = PatternFill("solid", fgColor="2E4053")  # cinza azulado — STATUS

    nova_price_display   = set(_price_short.values())
    antiga_price_display = set(antiga_col_headers)
    pct_display          = {f"% {lbl}" for lbl in _PRICE_LABELS}
    distri_display       = {"DISTRI NOVA", "DISTRI ANTIGA", "DISTRI %"} | set(_DC_LABELS)

    for ci, name in enumerate(final_cols, 1):
        cell = ws.cell(row=1, column=ci, value=name)
        if name in nova_price_display:
            cell.fill = _FILL_HDR_NOVA
        elif name in antiga_price_display:
            cell.fill = _FILL_HDR_ANTIGA
        elif name in pct_display:
            cell.fill = _FILL_HDR_PCT
        elif name in distri_display:
            cell.fill = _FILL_HDR_DISTRI
        elif name == "STATUS":
            cell.fill = _FILL_HDR_STATUS
        else:
            cell.fill = _FILL_HEADER
        cell.font      = _FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    status_pos = col_idx["STATUS"]

    # ── Dados ──────────────────────────────────────────────────────
    for wi, (df_idx, new_row) in enumerate(new_df.iterrows(), 2):
        info       = status_map.get(df_idx, {"status": "IGUAL", "pct": {}, "old_prices": {}})
        status     = info["status"]
        pct        = info.get("pct", {})
        old_prices = info.get("old_prices", {})

        # Valores NOVA (colunas originais renomeadas)
        for orig_name, display_name in price_orig_to_short.items():
            val  = new_row.get(orig_name)
            ci   = col_idx.get(display_name)
            if ci:
                cell           = ws.cell(row=wi, column=ci, value=_to_float(val) if val is not None else val)
                cell.alignment = Alignment(vertical="center")
                if status == "NOVO":
                    cell.fill = _FILL_NEW
                    cell.font = _FONT_NEW

        # Demais colunas originais (não-preço)
        # Colunas antes do ponto de inserção (cabeçalho original até 10+)
        for orig_name, display_name in zip(col_names[: insert_after + 1], display_col_names[: insert_after + 1]):
            if orig_name in price_orig_to_short:
                continue  # já tratado acima
            ci = col_idx.get(display_name)
            if ci:
                cell           = ws.cell(row=wi, column=ci, value=new_row.get(orig_name))
                cell.alignment = Alignment(vertical="center")
                if status == "NOVO":
                    cell.fill = _FILL_NEW
                    cell.font = _FONT_NEW
        # Colunas depois do ponto de inserção (filtradas para excluir colunas reservadas)
        for orig_name, display_name in zip(tail_col_names, tail_display_names):
            ci = col_idx.get(display_name)
            if ci:
                cell           = ws.cell(row=wi, column=ci, value=new_row.get(orig_name))
                cell.alignment = Alignment(vertical="center")
                if status == "NOVO":
                    cell.fill = _FILL_NEW
                    cell.font = _FONT_NEW


        # Colunas ANTIGA
        for lbl, hdr in zip(_PRICE_LABELS, antiga_col_headers):
            ci = col_idx.get(hdr)
            if ci:
                val  = old_prices.get(lbl)
                cell = ws.cell(row=wi, column=ci, value=val)
                cell.alignment = Alignment(vertical="center")

        # Colunas % variação
        for lbl in _PRICE_LABELS:
            ci      = col_idx.get(f"% {lbl}")
            pct_val = pct.get(lbl)
            if ci is None:
                continue
            cell               = ws.cell(row=wi, column=ci, value=pct_val)
            cell.alignment     = Alignment(horizontal="center", vertical="center")
            cell.number_format = "+0.00%;-0.00%;0.00%"
            if pct_val is not None:
                if pct_val > 1e-6:
                    cell.fill = _FILL_INC
                    cell.font = _FONT_INC
                elif pct_val < -1e-6:
                    cell.fill = _FILL_DEC
                    cell.font = _FONT_DEC

        # DISTRI NOVA = Preço 1-5 nova / 1.23
        price_15_new = _to_float(new_row.get(price_15_col_new)) if price_15_col_new else None
        distri_nova  = round(price_15_new / _DISTRI_DIVISOR, 10) if price_15_new else None
        ci_dn = col_idx.get("DISTRI NOVA")
        if ci_dn:
            cell           = ws.cell(row=wi, column=ci_dn, value=distri_nova)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Double-check ratios: preço_nova / DISTRI_NOVA
        dc_price_cols = {
            "DC 1-5":  new_cols.get("1-5"),
            "DC 6-10": new_cols.get("5-10"),
            "DC 10+":  new_cols.get("10+"),
        }
        for dc_lbl, price_col in dc_price_cols.items():
            ci = col_idx.get(dc_lbl)
            if ci is None:
                continue
            price_nova = _to_float(new_row.get(price_col)) if price_col else None
            ratio      = round(price_nova / distri_nova, 6) if (price_nova and distri_nova) else None
            cell               = ws.cell(row=wi, column=ci, value=ratio)
            cell.number_format = "0.0000"
            cell.alignment     = Alignment(horizontal="center", vertical="center")
            if ratio is not None:
                lims = _DC_LIMITS[dc_lbl]
                if ratio > lims["max"] or ratio < lims["min"]:
                    cell.fill = _FILL_DC_ALERT
                    cell.font = _FONT_DC_ALERT

        # DISTRI ANTIGA = Preço 1-5 antiga / 1.23
        price_15_old = old_prices.get("1-5")
        distri_antiga = round(price_15_old / _DISTRI_DIVISOR, 10) if price_15_old else None
        ci_da = col_idx.get("DISTRI ANTIGA")
        if ci_da:
            cell           = ws.cell(row=wi, column=ci_da, value=distri_antiga)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # DISTRI % = variação entre DISTRI NOVA e DISTRI ANTIGA
        ci_dpct = col_idx.get("DISTRI %")
        if ci_dpct:
            distri_pct = None
            if distri_nova is not None and distri_antiga:
                distri_pct = round((distri_nova - distri_antiga) / distri_antiga, 6)
            cell               = ws.cell(row=wi, column=ci_dpct, value=distri_pct)
            cell.number_format = "+0.00%;-0.00%;0.00%"
            cell.alignment     = Alignment(horizontal="center", vertical="center")
            if distri_pct is not None:
                if distri_pct > 1e-6:
                    cell.fill = _FILL_INC
                    cell.font = _FONT_INC
                elif distri_pct < -1e-6:
                    cell.fill = _FILL_DEC
                    cell.font = _FONT_DEC

        # STATUS
        sc           = ws.cell(row=wi, column=status_pos, value=status)
        sc.alignment = Alignment(horizontal="center", vertical="center")
        sc.font      = _FONT_STATUS

    # ── Largura das colunas ────────────────────────────────────────
    for ci, name in enumerate(final_cols, 1):
        w = max(len(str(name)) + 4, 12)
        ws.column_dimensions[get_column_letter(ci)].width = min(w, 48)

    # ── Linhas REMOVIDAS — adicionadas ao final da aba principal ──
    if removed_df is not None and len(removed_df) > 0:
        next_row      = ws.max_row + 1
        old_col_names = list(removed_df.columns)

        for _, rem_row in removed_df.iterrows():
            for col_name in final_cols:
                ci = col_idx.get(col_name)
                if ci is None:
                    continue
                if col_name == "STATUS":
                    val = "REMOVIDO"
                elif col_name in old_col_names:
                    val = rem_row[col_name]
                else:
                    # tenta casar prefixo "NOVA X" com coluna original X no removed_df
                    bare = col_name.replace("NOVA ", "").replace("ANTIGA ", "")
                    val  = rem_row.get(bare)

                cell           = ws.cell(row=next_row, column=ci, value=val)
                cell.fill      = _FILL_REMOVED
                cell.alignment = Alignment(vertical="center")
                cell.font      = (
                    Font(bold=True, italic=True, color="595959")
                    if col_name == "STATUS"
                    else _FONT_REMOVED
                )
                if col_name == "STATUS":
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            next_row += 1

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes    = "B2"   # congela até coluna B (Produto) + linha 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
